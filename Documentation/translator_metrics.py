import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os
from datetime import datetime, timedelta

# ─── LOAD DATA ────────────────────────────────────────────────────────────────

def load_data(filepath):
    sheets = pd.read_excel(filepath, sheet_name=None)
    print(f"Sheets found: {list(sheets.keys())}")
    return sheets

# ─── SCHEDULED WORKING HOURS BETWEEN TWO TIMESTAMPS ──────────────────────────

def build_schedule_lookup(schedules_df):
    """Build a dict: {name_lower: {day_int: (start_time, end_time)}}
    day_int: 0=Mon,1=Tue,2=Wed,3=Thu,4=Fri,5=Sat,6=Sun
    """
    day_cols = {0: 'MON', 1: 'TUES', 2: 'WED', 3: 'THURS', 4: 'FRI', 5: 'SAT', 6: 'SUN'}
    lookup = {}
    for _, row in schedules_df.iterrows():
        name = str(row['NAME']).strip().lower()
        try:
            # START/END may be time or string
            start = row['START']
            end = row['END']
            if hasattr(start, 'hour'):
                sh, sm = start.hour, start.minute
                eh, em = end.hour, end.minute
            else:
                parts = str(start).split(':')
                sh, sm = int(parts[0]), int(parts[1])
                parts = str(end).split(':')
                eh, em = int(parts[0]), int(parts[1])
        except Exception:
            sh, sm, eh, em = 8, 0, 18, 0

        avail = {}
        for d, col in day_cols.items():
            if col in row.index and row[col] == 1:
                avail[d] = (sh * 60 + sm, eh * 60 + em)  # minutes from midnight
        lookup[name] = avail
    return lookup

def scheduled_minutes(t_start, t_end, schedule):
    """Count actual working minutes between t_start and t_end using schedule.
    schedule: dict {weekday_int: (start_min, end_min)}
    """
    if pd.isnull(t_start) or pd.isnull(t_end):
        return np.nan
    if t_end <= t_start:
        return 0.0

    total = 0.0
    current = t_start
    while current.date() <= t_end.date():
        wd = current.weekday()
        if wd in schedule:
            day_start_min, day_end_min = schedule[wd]
            # Convert to datetime on this date
            ds = current.replace(hour=day_start_min // 60, minute=day_start_min % 60,
                                  second=0, microsecond=0)
            de = current.replace(hour=day_end_min // 60, minute=day_end_min % 60,
                                  second=0, microsecond=0)
            # Clip to actual window
            window_start = max(current if current.date() == t_start.date() else ds, ds)
            window_end = min(t_end if t_end.date() == current.date() else de, de)
            if current.date() == t_start.date():
                window_start = max(t_start, ds)
            if t_end.date() == current.date():
                window_end = min(t_end, de)
            if window_end > window_start:
                total += (window_end - window_start).total_seconds() / 60
        # Next day start of day
        current = (current + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)

    return total

# ─── METRICS COMPUTATION ──────────────────────────────────────────────────────

def compute_metrics(data_df, schedules_df):
    df = data_df.copy()

    # Normalise translator name
    df['_translator'] = df['TRANSLATOR'].astype(str).str.strip()

    # Parse timestamps
    for col in ['WORKING', 'DELIVERED']:
        df[col] = pd.to_datetime(df[col], errors='coerce')

    df['LANG_PAIR'] = df['SOURCE_LANG'].astype(str).str.strip() + ' → ' + df['TARGET_LANG'].astype(str).str.strip()

    # Build schedule lookup
    sched_lookup = build_schedule_lookup(schedules_df)

    # Compute actual scheduled minutes per row
    def row_actual_hours(row):
        name = row['_translator'].lower()
        sched = sched_lookup.get(name, {})
        if not sched:
            # Fallback: raw diff
            if pd.notnull(row['WORKING']) and pd.notnull(row['DELIVERED']):
                return (row['DELIVERED'] - row['WORKING']).total_seconds() / 3600
            return np.nan
        mins = scheduled_minutes(row['WORKING'], row['DELIVERED'], sched)
        return mins / 60 if not np.isnan(mins) else np.nan

    df['_actual_hours'] = df.apply(row_actual_hours, axis=1)
    df['_forecast_hours'] = pd.to_numeric(df['HOURS'], errors='coerce')
    df['_hours_diff'] = df['_actual_hours'] - df['_forecast_hours']

    # Quality
    df['_quality'] = pd.to_numeric(df['QUALITY_EVALUATION'], errors='coerce')

    # Sector column
    sector_col = None
    for c in df.columns:
        if 'manufacturer' in c.lower() and c.upper() != 'MANUFACTURER':
            sector_col = c
            break
        if 'sector' in c.lower():
            sector_col = c
            break
    if sector_col is None and 'MANUFACTURER_SECTOR' in df.columns:
        sector_col = 'MANUFACTURER_SECTOR'
    # Try last col starting with MANUFACTURER
    if sector_col is None:
        mfr_cols = [c for c in df.columns if str(c).upper().startswith('MANUFACTURER') and c != 'MANUFACTURER']
        if mfr_cols:
            sector_col = mfr_cols[-1]

    print(f"Sector column detected: {sector_col}")

    translators = sorted(df['_translator'].dropna().unique())

    # ── 1. Overall quality per translator ────────────────────────────────────
    overall_quality = (
        df.groupby('_translator')['_quality']
        .agg(Avg_Quality='mean', Task_Count='count', Quality_StdDev='std')
        .reset_index()
        .rename(columns={'_translator': 'Translator'})
    )
    overall_quality['Avg_Quality'] = overall_quality['Avg_Quality'].round(2)
    overall_quality['Quality_StdDev'] = overall_quality['Quality_StdDev'].round(2)

    # ── 2. Quality per translator per language pair ───────────────────────────
    lang_quality = (
        df.groupby(['_translator', 'LANG_PAIR'])['_quality']
        .agg(Avg_Quality='mean', Task_Count='count')
        .reset_index()
        .rename(columns={'_translator': 'Translator', 'LANG_PAIR': 'Language_Pair'})
    )
    lang_quality['Avg_Quality'] = lang_quality['Avg_Quality'].round(2)

    # ── 3. Forecast vs actual hours ───────────────────────────────────────────
    time_comparison = (
        df.groupby('_translator')
        .agg(
            Forecast_Hours=('_forecast_hours', 'sum'),
            Actual_Hours=('_actual_hours', 'sum'),
            Task_Count=('_hours_diff', 'count'),
            Avg_Diff_per_Task=('_hours_diff', 'mean')
        )
        .reset_index()
        .rename(columns={'_translator': 'Translator'})
    )
    time_comparison['Total_Diff_Hours'] = (time_comparison['Actual_Hours'] - time_comparison['Forecast_Hours']).round(2)
    time_comparison['Forecast_Hours'] = time_comparison['Forecast_Hours'].round(2)
    time_comparison['Actual_Hours'] = time_comparison['Actual_Hours'].round(2)
    time_comparison['Avg_Diff_per_Task'] = time_comparison['Avg_Diff_per_Task'].round(2)
    time_comparison['Efficiency_%'] = (
        (time_comparison['Forecast_Hours'] / time_comparison['Actual_Hours'].replace(0, np.nan)) * 100
    ).round(1)

    # ── 4. Sector experience ──────────────────────────────────────────────────
    if sector_col and sector_col in df.columns:
        sector_exp = (
            df.groupby(['_translator', sector_col])['_quality']
            .agg(Avg_Quality='mean', Task_Count='count')
            .reset_index()
            .rename(columns={'_translator': 'Translator', sector_col: 'Sector'})
        )
        sector_exp['Avg_Quality'] = sector_exp['Avg_Quality'].round(2)
        # Experience score: weighted by count and quality
        max_count = sector_exp['Task_Count'].max() or 1
        sector_exp['Experience_Score'] = (
            (sector_exp['Task_Count'] / max_count * 0.5 + sector_exp['Avg_Quality'] / 10 * 0.5) * 10
        ).round(2)
    else:
        sector_exp = pd.DataFrame(columns=['Translator', 'Sector', 'Avg_Quality', 'Task_Count', 'Experience_Score'])

    # ── 5. Task type experience ───────────────────────────────────────────────
    tasktype_exp = (
        df.groupby(['_translator', 'TASK_TYPE'])['_quality']
        .agg(Avg_Quality='mean', Task_Count='count')
        .reset_index()
        .rename(columns={'_translator': 'Translator', 'TASK_TYPE': 'Task_Type'})
    )
    tasktype_exp['Avg_Quality'] = tasktype_exp['Avg_Quality'].round(2)
    max_count_t = tasktype_exp['Task_Count'].max() or 1
    tasktype_exp['Experience_Score'] = (
        (tasktype_exp['Task_Count'] / max_count_t * 0.5 + tasktype_exp['Avg_Quality'] / 10 * 0.5) * 10
    ).round(2)

    return overall_quality, lang_quality, time_comparison, sector_exp, tasktype_exp

# ─── EXCEL STYLING ────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill('solid', start_color='1F3864')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
ALT_FILL    = PatternFill('solid', start_color='EEF2F7')
BODY_FONT   = Font(name='Arial', size=10)
BORDER      = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

def style_sheet(ws, df, title):
    # Title row
    ws.insert_rows(1)
    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, name='Arial', size=13, color='1F3864')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

    # Header row (row 2 after inserts)
    header_row = 2
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = col_name.replace('_', ' ')
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

    # Data rows
    for row_idx, (_, data_row) in enumerate(df.iterrows(), start=3):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for col_idx, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center' if col_idx > 1 else 'left')

    # Auto column widths
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max() if len(df) > 0 else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 35)

    ws.freeze_panes = ws.cell(row=3, column=1)
    ws.row_dimensions[2].height = 30

# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main(input_path, output_path):
    print(f"Loading: {input_path}")
    sheets = load_data(input_path)

    # Detect sheet names flexibly
    data_sheet = next((k for k in sheets if 'data' in k.lower()), list(sheets.keys())[0])
    sched_sheet = next((k for k in sheets if 'sched' in k.lower()), None)

    print(f"Using data sheet: '{data_sheet}', schedule sheet: '{sched_sheet}'")

    data_df = sheets[data_sheet]
    schedules_df = sheets[sched_sheet] if sched_sheet else pd.DataFrame()

    print(f"Data columns: {list(data_df.columns)}")
    print(f"Data rows: {len(data_df)}")

    overall_q, lang_q, time_cmp, sector_exp, task_exp = compute_metrics(data_df, schedules_df)

    # Write output
    wb = Workbook()
    wb.remove(wb.active)

    sheets_to_write = [
        ('1_Overall_Quality',   overall_q,   '1. Average Quality Evaluation per Translator'),
        ('2_Quality_by_LangPair', lang_q,    '2. Quality by Language Pair per Translator'),
        ('3_Time_Forecast_vs_Actual', time_cmp, '3. Forecast vs Actual Hours (Scheduled Working Time)'),
        ('4_Sector_Experience', sector_exp,  '4. Experience Score by Sector per Translator'),
        ('5_TaskType_Experience', task_exp,  '5. Experience Score by Task Type per Translator'),
    ]

    for sheet_name, df, title in sheets_to_write:
        ws = wb.create_sheet(title=sheet_name)
        # Write data first (openpyxl append)
        ws.append(list(df.columns))
        for _, row in df.iterrows():
            ws.append(list(row))
        style_sheet(ws, df, title)
        print(f"  ✓ Sheet '{sheet_name}' written ({len(df)} rows)")

    wb.save(output_path)
    print(f"\nSaved: {output_path}")

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: python translator_metrics.py <260319_GrauIAI_data.xlsx.xlsx> <New_data.xlsx>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
